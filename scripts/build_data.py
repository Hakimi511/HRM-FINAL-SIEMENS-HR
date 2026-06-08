# -*- coding: utf-8 -*-
"""
数据构建脚本 / Data build script
- 解析 25专家测试题.xlsx -> data/questions.json
- 复制并整理 PDF -> public/materials/ , 生成 data/materials.json
运行: python scripts/build_data.py
"""
import os, re, json, shutil
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC_XLSX = os.path.join(ROOT, '25专家测试题.xlsx')
DATA_DIR = os.path.join(ROOT, 'data')
MAT_DIR = os.path.join(ROOT, 'materials')
os.makedirs(DATA_DIR, exist_ok=True)

# ---------------------------------------------------------------------------
# 1) 模块归并：25个知识分类 -> 5大模块
# ---------------------------------------------------------------------------
MODULES = [
    {"id": "digital",     "name": "数字化产品",   "icon": "🔌", "desc": "西门子智能配电与能源数字化平台产品知识",
     "categories": ["数字化MGMS", "数字化NXpower", "Powermanager", "数字化ECX", "数字化SC Insights", "EA数字化SCADA", "EA数字化ISED"]},
    {"id": "sales",       "name": "销售技能",     "icon": "💼", "desc": "线索开发、解决方案销售与大客户管理",
     "categories": ["潜在客户生成", "发展竞争策略", "客户预测分析", "解决方案销售", "大客户管理", "销售专家的软技能", "武功秘籍"]},
    {"id": "negotiation", "name": "谈判与沟通",   "icon": "🤝", "desc": "高效谈判、沟通协作与多向管理",
     "categories": ["高效谈判的两大策略", "谈判技巧新思路训练", "跨智能沟通与协作技巧", "向上管理、向下管理与平行管理"]},
    {"id": "soft",        "name": "职场软技能",   "icon": "🧠", "desc": "情绪与冲突管理、抗挫力与时间管理",
     "categories": ["在工作中管理情绪", "冲突解决基础知识", "如何进行冲突管理", "如何正确面对销售挫折", "时间管理的法则与技巧", "时间管理与效率提升"]},
    {"id": "finance",     "name": "财务知识",     "icon": "📈", "desc": "企业财务报表分析",
     "categories": ["企业财务报表分析"]},
]
CAT2MODULE = {}
for m in MODULES:
    for c in m["categories"]:
        CAT2MODULE[c] = m["id"]

OPT_COLS = list("GHIJKLMNOP")  # 选项 1..10


def build_questions():
    wb = openpyxl.load_workbook(SRC_XLSX, data_only=True)
    ws = wb["Sheet"]
    questions = []
    unknown_cats = set()
    for r in range(2, ws.max_row + 1):
        def cell(col):
            return ws.cell(row=r, column=openpyxl.utils.column_index_from_string(col)).value
        qtext = cell('D')
        if qtext is None or str(qtext).strip() == "":
            continue
        cat = str(cell('C')).strip()
        module = CAT2MODULE.get(cat)
        if module is None:
            unknown_cats.add(cat)
            module = "other"
        # 选项
        options = []
        for col in OPT_COLS:
            v = cell(col)
            if v is not None and str(v).strip() != "":
                options.append(str(v).strip())
        # 答案：1-based -> 0-based
        raw_ans = str(cell('F')).strip()
        answer = []
        for a in re.split(r'[，,;；\s]+', raw_ans):
            a = a.strip()
            if a.isdigit():
                idx = int(a) - 1
                if 0 <= idx < len(options):
                    answer.append(idx)
        answer = sorted(set(answer))
        qtype = "multiple" if len(answer) > 1 else "single"
        expl = cell('T')
        expl = str(expl).strip() if expl is not None and str(expl).strip() != "" else None
        questions.append({
            "id": int(cell('B')) if cell('B') is not None else r,
            "category": cat,
            "module": module,
            "type": qtype,
            "question": str(qtext).strip(),
            "options": options,
            "answer": answer,
            "explanation": expl,
            "shuffle": str(cell('E')).strip().upper() == 'Y',
            "score": int(cell('Q')) if cell('Q') is not None else 1,
            "difficulty": int(cell('R')) if cell('R') is not None else 1,
        })
    # 分类统计
    cat_stats = {}
    for q in questions:
        cat_stats.setdefault(q["category"], {"module": q["module"], "count": 0, "single": 0, "multiple": 0})
        cat_stats[q["category"]]["count"] += 1
        cat_stats[q["category"]][q["type"]] += 1

    out = {
        "meta": {
            "title": "西门子 SI 渠道销售专家测试题库",
            "source": "25专家测试题.xlsx",
            "totalQuestions": len(questions),
            "totalCategories": len(cat_stats),
            "single": sum(1 for q in questions if q["type"] == "single"),
            "multiple": sum(1 for q in questions if q["type"] == "multiple"),
        },
        "modules": [{k: v for k, v in m.items()} for m in MODULES],
        "categories": [
            {"name": c, "module": s["module"], "count": s["count"], "single": s["single"], "multiple": s["multiple"]}
            for c, s in cat_stats.items()
        ],
        "questions": questions,
    }
    with open(os.path.join(DATA_DIR, 'questions.json'), 'w', encoding='utf-8') as f:
        json.dump(out, f, ensure_ascii=False, indent=1)
    print(f"[questions] {len(questions)} 题 / {len(cat_stats)} 分类 -> data/questions.json")
    if unknown_cats:
        print("  !! 未归类分类:", unknown_cats)


# ---------------------------------------------------------------------------
# 2) PDF 资料整理
# ---------------------------------------------------------------------------
# 学习资料分类的显式元数据（id / 徽标短码 / 中文描述），避免中文 slug 冲突
CAT_META = {
    "EA AIS 基础":  {"id": "ea-ais",       "code": "AIS",  "desc": "空气绝缘开关设备 (AIS)"},
    "EA CS 基础":   {"id": "ea-cs",        "code": "CS",   "desc": "业务服务方案与销售政策"},
    "EA GIS 基础":  {"id": "ea-gis",       "code": "GIS",  "desc": "气体绝缘开关设备 (GIS)"},
    "EA 保护 基础": {"id": "ea-protection","code": "保护", "desc": "保护与测控装置"},
    "EA 数字化 基础":{"id": "ea-digital",  "code": "数字化","desc": "智能配电数字化解决方案"},
    "VCB 基础":     {"id": "ea-vcb",       "code": "VCB",  "desc": "真空断路器 (VCB)"},
    "8PT 基础":     {"id": "ep-8pt",       "code": "8PT",  "desc": "Sivacon 8PT 低压成套开关设备"},
    "ACB 基础":     {"id": "ep-acb",       "code": "ACB",  "desc": "万能式空气断路器 (ACB)"},
    "ATSE LBS 基础":{"id": "ep-atse-lbs",  "code": "ATSE", "desc": "双电源转换开关与负荷开关"},
    "Digital 基础": {"id": "ep-digital",   "code": "Digital","desc": "Panel Manager 数字化电柜管家"},
    "MCB RCD基础":  {"id": "ep-mcb-rcd",   "code": "MCB",  "desc": "微型断路器与剩余电流保护"},
    "MCCB 基础":    {"id": "ep-mccb",      "code": "MCCB", "desc": "塑壳断路器 (MCCB)"},
}
LINE_INFO = {
    "EA 产品": {"id": "ea", "name": "EA 产品线（电气化与自动化 · 中压）", "icon": "⚡",
                "desc": "中压开关设备、保护测控与配电数字化"},
    "EP 产品": {"id": "ep", "name": "EP 产品线（电气产品 · 低压）", "icon": "🔋",
                "desc": "低压断路器、成套设备与电柜数字化"},
}


def slugify(name):
    s = name.lower()
    s = re.sub(r'[^a-z0-9]+', '-', s)
    s = s.strip('-')
    return s


def build_materials():
    if os.path.exists(MAT_DIR):
        shutil.rmtree(MAT_DIR)
    os.makedirs(MAT_DIR, exist_ok=True)

    lines = []
    total_files = 0
    for line_folder in sorted(os.listdir(ROOT)):
        lp = os.path.join(ROOT, line_folder)
        if not os.path.isdir(lp) or line_folder not in LINE_INFO:
            continue
        info = LINE_INFO[line_folder]
        line_id = info["id"]
        categories = []
        for cat_folder in sorted(os.listdir(lp)):
            cp = os.path.join(lp, cat_folder)
            if not os.path.isdir(cp):
                continue
            meta = CAT_META.get(cat_folder, {"id": line_id + '-' + slugify(cat_folder), "code": "", "desc": ""})
            cat_id = meta["id"]
            files = []
            pdfs = sorted([f for f in os.listdir(cp) if f.lower().endswith('.pdf')])
            for i, fn in enumerate(pdfs, 1):
                title = os.path.splitext(fn)[0]
                slug = slugify(title) or str(i)
                dest_rel = f"materials/{line_id}/{cat_id}-{slug}.pdf"
                dest_abs = os.path.join(ROOT, dest_rel.replace('/', os.sep))
                os.makedirs(os.path.dirname(dest_abs), exist_ok=True)
                shutil.copyfile(os.path.join(cp, fn), dest_abs)
                size = os.path.getsize(dest_abs)
                files.append({"title": title, "file": dest_rel, "sizeMB": round(size / 1024 / 1024, 2)})
                total_files += 1
            categories.append({
                "id": cat_id,
                "name": cat_folder,
                "code": meta["code"],
                "desc": meta["desc"],
                "files": files,
            })
        lines.append({
            "id": line_id, "name": info["name"], "icon": info["icon"], "desc": info["desc"],
            "categories": categories,
        })

    out = {
        "meta": {
            "title": "西门子 SI 产品学习资料库",
            "totalLines": len(lines),
            "totalCategories": sum(len(l["categories"]) for l in lines),
            "totalFiles": total_files,
        },
        "lines": lines,
    }
    with open(os.path.join(DATA_DIR, 'materials.json'), 'w', encoding='utf-8') as f:
        json.dump(out, f, ensure_ascii=False, indent=1)
    print(f"[materials] {total_files} 个PDF -> materials/ , data/materials.json")


if __name__ == '__main__':
    build_questions()
    build_materials()
    print("DONE")
